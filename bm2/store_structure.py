from __future__ import annotations

from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .constants import (
    EXPENSE_META_HEADERS,
    EXPENSE_META_SHEET,
    EXPENSE_META_SHEET_ALIASES,
    EXPENSE_NAME_HEADER,
    INCOME_META_HEADERS,
    INCOME_META_SHEET,
    INCOME_META_SHEET_ALIASES,
    INCOME_NAME_HEADER,
    META_HEADERS,
    META_SHEET,
    META_SHEET_ALIASES,
    NAME_HEADER,
    PROFIT_HEADERS,
    PROFIT_HEADER,
    PROFIT_SHEET,
    PROFIT_SHEET_ALIASES,
    SCORE_SHEET,
    TOTAL_HEADER,
    WEAR_META_HEADERS,
    LEGACY_WEAR_HEADERS,
    OLD_SCORE_PREFIX,
    WEAR_META_SHEET,
    WEAR_META_SHEET_ALIASES,
    WEAR_NAME_HEADER,
    WEAR_SHEET,
    WEAR_TOTAL_HEADER,
    WINDOW_SIZE,
)


class StoreStructureMixin:
    def _ensure_sheet_member_rows(self, sheet, *, name_col: int, total_col: int | None, value_columns: list[int], extra_columns: list[int] | None = None) -> bool:
        _, changed = self._ensure_member_rows(
            sheet,
            name_col=name_col,
            total_col=total_col,
            value_columns=value_columns,
            extra_columns=extra_columns,
        )
        return changed

    def _finalize_score_sheet_structure(self, workbook, sheet, *, total_col: int, profit_col: int, name_col: int) -> None:
        self._recalculate_totals(sheet, total_col)
        self._recalculate_score_profits(workbook, sheet, profit_col)
        self._sort_named_rows(sheet, total_col, name_col)
        recent_numbers = [number for number, _ in self._d_columns(sheet)][-WINDOW_SIZE:]
        self._format_score_sheet(sheet, recent_numbers, total_col)

    def _finalize_wear_sheet_structure(self, sheet, *, total_col: int, name_col: int) -> None:
        self._recalculate_wear_totals(sheet, total_col)
        self._sort_named_rows(sheet, total_col, name_col)
        self._format_wear_sheet(sheet, total_col)

    def _finalize_value_sheet_structure(self, sheet, *, name_col: int) -> None:
        self._sort_rows_by_name(sheet, name_col)

    def _create_initial_score_sheet(self, workbook) -> None:
        score_sheet = workbook.active
        score_sheet.title = SCORE_SHEET
        initial_header_count = 41
        end_date = datetime.now().date()
        for column_index in range(1, initial_header_count + 1):
            header_date = end_date - timedelta(days=(initial_header_count - column_index))
            score_sheet.cell(1, column_index, header_date.strftime('%m-%d'))
        score_sheet.cell(1, initial_header_count + 1, TOTAL_HEADER)
        score_sheet.cell(1, initial_header_count + 2, NAME_HEADER)

    def _ensure_meta_sheets(self, workbook) -> bool:
        changed = False
        _, meta_changed = self._ensure_named_sheet(workbook, META_SHEET, META_SHEET_ALIASES, META_HEADERS, hidden=True)
        changed = meta_changed or changed
        _, wear_meta_changed = self._ensure_named_sheet(workbook, WEAR_META_SHEET, WEAR_META_SHEET_ALIASES, WEAR_META_HEADERS, hidden=True)
        changed = wear_meta_changed or changed
        _, income_meta_changed = self._ensure_named_sheet(workbook, INCOME_META_SHEET, INCOME_META_SHEET_ALIASES, INCOME_META_HEADERS, hidden=True)
        changed = income_meta_changed or changed
        _, expense_meta_changed = self._ensure_named_sheet(workbook, EXPENSE_META_SHEET, EXPENSE_META_SHEET_ALIASES, EXPENSE_META_HEADERS, hidden=True)
        changed = expense_meta_changed or changed
        return changed

    def _ensure_score_sheet_structure(self, workbook) -> bool:
        sheet = self._score_sheet(workbook)
        changed = False

        col = 1
        while col <= sheet.max_column:
            if sheet.cell(1, col).value is not None:
                col += 1
                continue
            if any(sheet.cell(row, col).value is not None for row in range(2, sheet.max_row + 1)):
                col += 1
                continue
            sheet.delete_cols(col)
            changed = True

        total_col, profit_col, name_col, tail_changed = self._ensure_score_tail_columns(sheet)
        changed = changed or tail_changed
        changed = self._repair_score_headers(sheet, total_col, profit_col) or changed

        changed = self._ensure_sheet_member_rows(
            sheet,
            name_col=name_col,
            total_col=total_col,
            value_columns=[col for _, col in self._d_columns(sheet)],
            extra_columns=[profit_col],
        ) or changed

        self._finalize_score_sheet_structure(workbook, sheet, total_col=total_col, profit_col=profit_col, name_col=name_col)
        return changed

    def _repair_score_headers(self, sheet, total_col: int, profit_col: int) -> bool:
        score_columns = [col for _, col in self._d_columns(sheet)]
        if not score_columns:
            return False
        headers = [sheet.cell(1, col).value for col in score_columns]
        has_legacy_header = any(self._parse_d_header(value) is not None for value in headers)
        has_mmdd_headers = all(self._is_mmdd_header(value) for value in headers if value is not None)
        if not has_legacy_header and not has_mmdd_headers:
            return False

        end_date = datetime.now().date()
        expected_headers = [
            (end_date - timedelta(days=(len(score_columns) - index - 1))).strftime('%m-%d')
            for index in range(len(score_columns))
        ]
        if headers == expected_headers:
            return False

        for header_value, col in zip(expected_headers, score_columns):
            sheet.cell(1, col, header_value)
        sheet.cell(1, total_col, TOTAL_HEADER)
        sheet.cell(1, profit_col, PROFIT_HEADER)
        sheet.cell(1, total_col + 2, NAME_HEADER)
        return True


    def _recalculate_score_profits(self, workbook, score_sheet, profit_col: int) -> None:
        score_name_col = self._find_column(score_sheet, NAME_HEADER)
        income_sheet = self._income_sheet(workbook)
        wear_sheet = self._wear_sheet(workbook)
        income_name_col = self._find_column(income_sheet, INCOME_NAME_HEADER)
        wear_name_col = self._find_column(wear_sheet, WEAR_NAME_HEADER)
        if score_name_col is None or income_name_col is None or wear_name_col is None:
            return

        income_row_map = self._build_name_row_map(income_sheet, income_name_col)
        wear_row_map = self._build_name_row_map(wear_sheet, wear_name_col)
        income_columns = [col for _, col in self._income_columns(income_sheet)]
        wear_columns = [col for _, col in self._wear_columns(wear_sheet)]

        for row in range(2, score_sheet.max_row + 1):
            member_name = str(score_sheet.cell(row, score_name_col).value or '').strip()
            if not member_name:
                continue
            income_total = self._round_income(
                self._sum_sheet_row_values(income_sheet, income_row_map[member_name], income_columns)
                if member_name in income_row_map else 0
            )
            wear_total = self._round_wear(
                self._sum_sheet_row_values(wear_sheet, wear_row_map[member_name], wear_columns)
                if member_name in wear_row_map else 0
            )
            score_sheet.cell(row, profit_col, self._round_income(income_total - wear_total))

    def _recalculate_wear_totals(self, sheet, total_col: int) -> None:
        name_col = self._find_column(sheet, WEAR_NAME_HEADER)
        if name_col is None:
            return
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row, name_col).value:
                continue
            total = 0.0
            for _, col in self._wear_columns(sheet):
                value = sheet.cell(row, col).value
                numeric = self._round_wear(value or 0)
                sheet.cell(row, col, numeric)
                total += numeric
            sheet.cell(row, total_col, self._round_wear(total))

    def _format_wear_sheet(self, sheet, total_col: int) -> None:
        threshold = self.get_wear_abnormal_threshold()
        normal_font = Font(color='000000', bold=False)
        abnormal_font = Font(color='FF0000', bold=True)
        for _, wear_col in self._wear_columns(sheet):
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row, wear_col)
                value = cell.value
                if value in (None, ''):
                    cell.font = normal_font
                    continue
                try:
                    numeric_value = float(value)
                except (TypeError, ValueError):
                    cell.font = normal_font
                    continue
                cell.font = abnormal_font if numeric_value > threshold else normal_font

        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, total_col).font = normal_font

    def _ensure_wear_sheet_structure(self, workbook) -> bool:
        sheet = self._wear_sheet(workbook)
        changed = False

        legacy_headers = ["日期", "姓名", "每日磨损"]
        current_headers = [sheet.cell(1, idx).value for idx in range(1, 4)]
        if current_headers == legacy_headers:
            self._migrate_legacy_wear_rows(workbook, sheet)
            changed = True

        total_col, name_col, tail_changed = self._ensure_summary_columns(sheet, WEAR_TOTAL_HEADER, WEAR_NAME_HEADER)
        changed = changed or tail_changed

        changed = self._ensure_sheet_member_rows(
            sheet,
            name_col=name_col,
            total_col=total_col,
            value_columns=[col for _, col in self._wear_columns(sheet)],
        ) or changed

        self._finalize_wear_sheet_structure(sheet, total_col=total_col, name_col=name_col)
        return changed

    def _ensure_income_sheet_structure(self, workbook) -> bool:
        return self._ensure_value_sheet_structure(
            workbook,
            sheet_getter=self._income_sheet,
            name_header=INCOME_NAME_HEADER,
            columns_getter=self._income_columns,
        )

    def _ensure_expense_sheet_structure(self, workbook) -> bool:
        return self._ensure_value_sheet_structure(
            workbook,
            sheet_getter=self._expense_sheet,
            name_header=EXPENSE_NAME_HEADER,
            columns_getter=self._expense_columns,
        )


    def _safe_float(self, value) -> float:
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

    def _sum_sheet_row_values(self, sheet, row: int, columns: list[int]) -> float:
        return sum(self._safe_float(sheet.cell(row, col).value) for col in columns)

    def _sort_profit_rows(self, sheet, *, name_col: int, profit_col: int) -> None:
        rows = []
        for row in range(2, sheet.max_row + 1):
            values = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
            if values[name_col - 1]:
                rows.append(values)
        rows.sort(key=lambda row: (-self._safe_float(row[profit_col - 1]), str(row[name_col - 1])))
        for row_index, values in enumerate(rows, start=2):
            for col_index, value in enumerate(values, start=1):
                sheet.cell(row_index, col_index, value)

    def _recalculate_profit_sheet(self, profit_sheet, income_sheet, wear_sheet) -> None:
        profit_name_col = self._find_column(profit_sheet, NAME_HEADER)
        income_name_col = self._find_column(income_sheet, INCOME_NAME_HEADER)
        wear_name_col = self._find_column(wear_sheet, WEAR_NAME_HEADER)
        if profit_name_col is None or income_name_col is None or wear_name_col is None:
            return

        income_row_map = self._build_name_row_map(income_sheet, income_name_col)
        wear_row_map = self._build_name_row_map(wear_sheet, wear_name_col)
        income_columns = [col for _, col in self._income_columns(income_sheet)]
        wear_columns = [col for _, col in self._wear_columns(wear_sheet)]

        for row in range(2, profit_sheet.max_row + 1):
            member_name = str(profit_sheet.cell(row, profit_name_col).value or '').strip()
            if not member_name:
                continue
            income_total = self._round_income(
                self._sum_sheet_row_values(income_sheet, income_row_map[member_name], income_columns)
                if member_name in income_row_map else 0
            )
            wear_total = self._round_wear(
                self._sum_sheet_row_values(wear_sheet, wear_row_map[member_name], wear_columns)
                if member_name in wear_row_map else 0
            )
            profit_sheet.cell(row, 2, income_total)
            profit_sheet.cell(row, 3, wear_total)
            profit_sheet.cell(row, 4, self._round_income(income_total - wear_total))

    def _ensure_profit_sheet_structure(self, workbook) -> bool:
        sheet, changed = self._ensure_named_sheet(workbook, PROFIT_SHEET, PROFIT_SHEET_ALIASES, PROFIT_HEADERS)
        name_col = self._find_column(sheet, NAME_HEADER)
        if name_col is None:
            return changed
        changed = self._ensure_sheet_member_rows(sheet, name_col=name_col, total_col=None, value_columns=[2, 3, 4]) or changed
        self._recalculate_profit_sheet(sheet, self._income_sheet(workbook), self._wear_sheet(workbook))
        self._sort_profit_rows(sheet, name_col=name_col, profit_col=4)
        return changed

    def _ensure_value_sheet_structure(self, workbook, *, sheet_getter, name_header: str, columns_getter) -> bool:
        sheet = sheet_getter(workbook)
        changed = False

        name_col = self._find_column(sheet, name_header)
        if name_col is None:
            sheet.cell(1, sheet.max_column + 1, name_header)
            changed = True
            name_col = self._find_column(sheet, name_header)

        assert name_col is not None
        if name_col != sheet.max_column:
            name_values = [sheet.cell(row, name_col).value for row in range(1, sheet.max_row + 1)]
            sheet.delete_cols(name_col)
            insert_at = sheet.max_column + 1
            sheet.insert_cols(insert_at, 1)
            for row, value in enumerate(name_values, start=1):
                sheet.cell(row, insert_at, value)
            changed = True
            name_col = insert_at

        changed = self._ensure_sheet_member_rows(
            sheet,
            name_col=name_col,
            total_col=None,
            value_columns=[col for _, col in columns_getter(sheet)],
        ) or changed
        self._finalize_value_sheet_structure(sheet, name_col=name_col)
        return changed

    # 首次启动时补齐工作簿和隐藏辅助页。
    def _ensure_workbook(self) -> None:
        changed = False
        if self.workbook_path.exists():
            workbook = load_workbook(self.workbook_path)
        else:
            workbook = Workbook()
            self._create_initial_score_sheet(workbook)
            changed = True

        changed = self._ensure_meta_sheets(workbook) or changed
        changed = self._ensure_score_sheet_structure(workbook) or changed
        changed = self._ensure_wear_sheet_structure(workbook) or changed
        changed = self._ensure_income_sheet_structure(workbook) or changed
        changed = self._ensure_expense_sheet_structure(workbook) or changed
        if PROFIT_SHEET in workbook.sheetnames:
            workbook.remove(workbook[PROFIT_SHEET])
            changed = True
        if changed:
            self._save_workbook(workbook)
        workbook.close()

    def _append_meta(self, workbook, date_text: str, col_name: str) -> None:
        workbook[META_SHEET].append([date_text, col_name])

    def _append_wear_meta(self, workbook, date_text: str, col_name: str) -> None:
        workbook[WEAR_META_SHEET].append([date_text, col_name])

    def _append_income_meta(self, workbook, date_text: str, col_name: str) -> None:
        workbook[INCOME_META_SHEET].append([date_text, col_name])

    def _append_expense_meta(self, workbook, date_text: str, col_name: str) -> None:
        workbook[EXPENSE_META_SHEET].append([date_text, col_name])

    def _read_sheet_meta(self, workbook, sheet_name: str) -> list[tuple[str, str]]:
        if sheet_name not in workbook.sheetnames:
            return []
        rows = []
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                rows.append((str(row[0]), str(row[1])))
        return rows

    def _normalize_wear_date_text(self, value: str, year: int) -> str:
        if len(value) == 4 and value.isdigit():
            return f"{year:04d}-{value[:2]}-{value[2:]}"
        return value

    def _migrate_legacy_wear_rows(self, workbook, sheet) -> None:
        legacy_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            legacy_rows.append(
                {
                    'date': str(row[0]),
                    'name': str(row[1]),
                    'wear': float(row[2] or 0),
                }
            )

        sheet.delete_rows(1, sheet.max_row)
        meta_sheet = workbook[WEAR_META_SHEET] if WEAR_META_SHEET in workbook.sheetnames else workbook.create_sheet(WEAR_META_SHEET)
        meta_sheet.delete_rows(1, meta_sheet.max_row)
        meta_sheet.append(WEAR_META_HEADERS)

        date_order: list[str] = []
        member_names: list[str] = []
        data_map: dict[tuple[str, str], float] = {}
        for item in legacy_rows:
            date_text = item['date']
            member_name = item['name']
            if date_text not in date_order:
                date_order.append(date_text)
            if member_name not in member_names:
                member_names.append(member_name)
            data_map[(member_name, date_text)] = item['wear']

        for index, date_text in enumerate(date_order, start=1):
            sheet.cell(1, index, date_text[5:].replace('-', ''))
            meta_sheet.append([date_text, str(index)])

        total_col = len(date_order) + 1
        name_col = total_col + 1
        sheet.cell(1, total_col, WEAR_TOTAL_HEADER)
        sheet.cell(1, name_col, WEAR_NAME_HEADER)
        for member_name in member_names:
            row = sheet.max_row + 1
            running_total = 0.0
            for index, date_text in enumerate(date_order, start=1):
                wear = self._round_wear(data_map.get((member_name, date_text), 0.0))
                sheet.cell(row, index, wear)
                running_total += wear
            sheet.cell(row, total_col, self._round_wear(running_total))
            sheet.cell(row, name_col, member_name)

    def _recalculate_totals(self, sheet, total_col: int) -> None:
        name_col = self._find_column(sheet, NAME_HEADER)
        if name_col is None:
            return
        d_cols = self._d_columns(sheet)
        recent_numbers = {number for number, _ in d_cols[-WINDOW_SIZE:]}
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row, name_col).value:
                continue
            total = 0
            for number, d_col in d_cols:
                value = sheet.cell(row, d_col).value
                numeric = int(value or 0)
                sheet.cell(row, d_col, numeric)
                if number in recent_numbers:
                    total += numeric
            sheet.cell(row, total_col, total)

    def _format_score_sheet(self, sheet, recent_numbers: list[int], total_col: int) -> None:
        old_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        light = (204, 255, 204)
        dark = (0, 170, 0)
        recent_set = set(recent_numbers)
        for number, d_col in self._d_columns(sheet):
            is_recent = number in recent_set
            fill = PatternFill(fill_type=None) if is_recent else old_fill
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, d_col).fill = fill

        totals = []
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row, total_col).value
            if isinstance(value, (int, float)):
                totals.append(float(value))
        max_score = max(totals) if totals else 0.0
        min_score = min(totals) if totals else 0.0
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row, total_col).value
            score = float(value) if isinstance(value, (int, float)) else 0.0
            if max_score == min_score:
                color = '99FF99'
            else:
                ratio = (score - min_score) / (max_score - min_score)
                red = int(light[0] + (dark[0] - light[0]) * ratio)
                green = int(light[1] + (dark[1] - light[1]) * ratio)
                blue = int(light[2] + (dark[2] - light[2]) * ratio)
                color = f'{red:02X}{green:02X}{blue:02X}'
            sheet.cell(row, total_col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
